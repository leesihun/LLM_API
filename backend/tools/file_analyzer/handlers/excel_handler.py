"""
Excel File Handler
===================
Handler for analyzing Excel files (.xlsx, .xls, .xlsm) with advanced metadata extraction.

Version: 1.0.0
Created: 2025-01-13
"""

from typing import Dict, Any, List

from backend.utils.logging_utils import get_logger
from ..base_handler import BaseFileHandler

logger = get_logger(__name__)


class ExcelHandler(BaseFileHandler):
    """
    Handler for Excel file analysis.

    Features:
    - Multi-sheet analysis
    - Formula detection and extraction
    - Named ranges identification
    - Merged cells detection
    - Data quality metrics per sheet
    - Comprehensive metadata extraction
    """

    def __init__(self):
        """Initialize Excel handler."""
        self.supported_extensions = ['xlsx', 'xls', 'xlsm']

    def supports(self, file_path: str) -> bool:
        """
        Check if this is an Excel file.

        Args:
            file_path: Path to the file

        Returns:
            True if file has Excel extension
        """
        extension = self.get_file_extension(file_path)
        return extension in self.supported_extensions

    def get_supported_extensions(self) -> List[str]:
        """Get list of supported extensions."""
        return self.supported_extensions

    def analyze(self, file_path: str) -> Dict[str, Any]:
        """
        Analyze Excel file with advanced metadata extraction.

        Args:
            file_path: Path to the Excel file

        Returns:
            Dictionary with analysis results including:
            - format: 'Excel'
            - total_sheets: Number of sheets
            - sheet_names: List of sheet names
            - sheets_analyzed: Detailed analysis per sheet
            - has_formulas: Boolean indicating formula presence
            - formulas: List of detected formulas
            - has_named_ranges: Boolean indicating named ranges
            - named_ranges: List of named ranges
            - has_merged_cells: Boolean indicating merged cells
            - merged_cells: List of merged cell ranges
        """
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter

            # Get all sheet names
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names

            # Load workbook for advanced metadata
            wb = None
            try:
                wb = load_workbook(file_path, data_only=False, read_only=True)
            except Exception:
                pass  # Continue without advanced features

            sheets_info = []
            formulas_found = []
            named_ranges = []
            merged_cells_info = []

            # Analyze ALL sheets
            for sheet_name in sheet_names:
                sheet_info = self._analyze_sheet(
                    file_path, sheet_name, wb, formulas_found, merged_cells_info
                )
                sheets_info.append(sheet_info)

            # Extract workbook-level metadata
            if wb:
                named_ranges = self._extract_named_ranges(wb)

            return {
                "format": "Excel",
                "total_sheets": len(sheet_names),
                "sheet_names": sheet_names,
                "sheets_analyzed": sheets_info,
                "has_formulas": len(formulas_found) > 0,
                "formulas": formulas_found if formulas_found else None,
                "has_named_ranges": len(named_ranges) > 0,
                "named_ranges": named_ranges[:20] if named_ranges else None,
                "has_merged_cells": len(merged_cells_info) > 0,
                "merged_cells": merged_cells_info if merged_cells_info else None
            }

        except ImportError:
            return {"format": "Excel", "error": "pandas or openpyxl not installed"}
        except Exception as e:
            logger.error(f"Excel analysis error: {e}", exc_info=True)
            return {"format": "Excel", "error": str(e)}

    def _analyze_sheet(
        self, file_path: str, sheet_name: str, wb, formulas_found: List, merged_cells_info: List
    ) -> Dict[str, Any]:
        """
        Analyze a single Excel sheet.

        Args:
            file_path: Path to Excel file
            sheet_name: Name of sheet to analyze
            wb: Openpyxl workbook object (or None)
            formulas_found: List to append found formulas
            merged_cells_info: List to append merged cell info

        Returns:
            Dictionary with sheet analysis results
        """
        import pandas as pd
        from openpyxl.utils import get_column_letter

        # Use pandas for data analysis
        df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=100)
        full_df = pd.read_excel(file_path, sheet_name=sheet_name)

        # Data quality metrics
        null_counts = full_df.isnull().sum().to_dict()
        total_nulls = sum(null_counts.values())
        duplicate_rows = full_df.duplicated().sum()

        sheet_info = {
            "sheet_name": sheet_name,
            "rows": len(full_df),
            "columns": len(full_df.columns),
            "column_names": list(full_df.columns),
            "column_types": {col: str(dtype) for col, dtype in full_df.dtypes.items()},
            "preview": df.head(3).to_dict(orient='records'),
            "null_counts": null_counts,
            "total_null_values": int(total_nulls),
            "duplicate_rows": int(duplicate_rows)
        }

        # Extract advanced metadata using openpyxl
        if wb:
            try:
                ws = wb[sheet_name]

                # Find formulas (sample first 100 cells)
                formulas_in_sheet = []
                for row in list(ws.iter_rows(max_row=min(100, ws.max_row), max_col=ws.max_column))[:20]:
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            formulas_in_sheet.append({
                                "cell": f"{get_column_letter(cell.column)}{cell.row}",
                                "formula": str(cell.value)[:100]
                            })

                if formulas_in_sheet:
                    formulas_found.append({
                        "sheet": sheet_name,
                        "formulas": formulas_in_sheet[:10]
                    })

                # Detect merged cells
                if hasattr(ws, 'merged_cells') and ws.merged_cells:
                    merged_ranges = [str(merged_cell) for merged_cell in list(ws.merged_cells.ranges)[:10]]
                    if merged_ranges:
                        merged_cells_info.append({
                            "sheet": sheet_name,
                            "merged_ranges": merged_ranges
                        })

                sheet_info["has_formulas"] = len(formulas_in_sheet) > 0
                sheet_info["formula_count"] = len(formulas_in_sheet)
                sheet_info["has_merged_cells"] = len(ws.merged_cells) > 0 if hasattr(ws, 'merged_cells') else False

            except Exception as e:
                logger.debug(f"Could not extract advanced metadata from sheet {sheet_name}: {e}")

        return sheet_info

    def _extract_named_ranges(self, wb) -> List[Dict[str, Any]]:
        """
        Extract named ranges from workbook.

        Args:
            wb: Openpyxl workbook object

        Returns:
            List of named range dictionaries
        """
        named_ranges = []

        try:
            if hasattr(wb, 'defined_names'):
                for named_range in wb.defined_names.definedName:
                    named_ranges.append({
                        "name": named_range.name,
                        "scope": named_range.localSheetId if named_range.localSheetId is not None else "Workbook",
                        "refers_to": str(named_range.attr_text)[:100]
                    })
        except Exception as e:
            logger.debug(f"Could not extract named ranges: {e}")

        return named_ranges
